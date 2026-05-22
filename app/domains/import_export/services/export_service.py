"""
KAiTix — Export Service
Ablegen: app/services/export_service.py

Sheets: Übersicht | Rack-Belegung | Ports & Interfaces | Kabelliste | PDU-Belegung | Import-Vorlage
Formate: xlsx, ods, csv (als ZIP)
"""

import csv
import io
import zipfile
from datetime import datetime
from typing import Literal

COL_HEADER = "1A1A2E"
COL_SERVER = "D6E4F7"
COL_SWITCH = "D6F0E8"
COL_PDU = "FFF3CD"
COL_FIREWALL = "FFE0D6"
COL_STORAGE = "EDE0F7"
COL_XRACK = "FFD6C0"
COL_WARN = "FFC7C7"
COL_WHITE = "FFFFFF"
COL_ALT = "F5F5F5"

DEVICE_COLOR = {
    "server": COL_SERVER,
    "switch": COL_SWITCH,
    "pdu": COL_PDU,
    "firewall": COL_FIREWALL,
    "storage": COL_STORAGE,
    "sonstige": COL_ALT,
}
PHASE_COL = {"L1": "D6EAF8", "L2": "D5F5E3", "L3": "FDEBD0"}


def _dev_map(data):
    return {d["id"]: d for d in data["devices"]}


def _rack_map(data):
    return {r["id"]: r for r in data["racks"]}


def _phase_loads(data, rack_id):
    ph = {"L1": 0.0, "L2": 0.0, "L3": 0.0}
    for d in data["devices"]:
        if d.get("rack_id") == rack_id and d.get("phase") in ph:
            ph[d["phase"]] += d.get("watt", 0)
    return ph


def _rack_devices(data, rack_id):
    return sorted(
        [d for d in data["devices"] if d.get("rack_id") == rack_id],
        key=lambda d: -(d.get("u_pos") or 0),
    )


def _ifaces_for_device(data, dev_id):
    return [i for i in data.get("interfaces", []) if i.get("dev_id") == dev_id]


def _cable_for_iface(data, kabel_id):
    if not kabel_id:
        return None
    return next((c for c in data.get("cables", []) if c["id"] == kabel_id), None)


def _pdu_outlets(data, pdu_id):
    return [o for o in data.get("pdu_outlets", []) if o.get("pdu_id") == pdu_id]


def _xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hf():
        return Font(bold=True, color="FFFFFF", name="Arial", size=10)

    def hfill(c):
        return PatternFill("solid", fgColor=c)

    def cfill(c):
        return PatternFill("solid", fgColor=c)

    def aln(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=True)

    def write_header(ws, cols, col_hex=COL_HEADER):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = hf()
            cell.fill = hfill(col_hex)
            cell.border = brd
            cell.alignment = aln("center")
        ws.row_dimensions[1].height = 20

    def style_row(ws, row_idx, fill_hex=None, bold=False):
        fill = cfill(fill_hex) if fill_hex else None
        for cell in ws[row_idx]:
            cell.font = Font(bold=bold, name="Arial", size=9)
            cell.border = brd
            cell.alignment = aln()
            if fill:
                cell.fill = fill

    def autowidth(ws, mn=8, mx=40):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(w + 2, mn), mx
            )

    ts = datetime.now().strftime("%d.%m.%Y %H:%M")
    dm = _dev_map(data)
    rm = _rack_map(data)

    # Übersicht
    ws = wb.create_sheet("Übersicht")
    ws["A1"] = "KAiTix — Technische Dokumentation"
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws["A2"] = f"Erstellt: {ts}"
    ws["A2"].font = Font(size=9, name="Arial", color="888888")
    ws.append([])
    write_header(
        ws, ["Rack", "Standort", "HE", "Geräte", "L1 kW", "L2 kW", "L3 kW", "Gesamt kW"]
    )
    for r in data["racks"]:
        ph = _phase_loads(data, r["id"])
        devs = _rack_devices(data, r["id"])
        ws.append(
            [
                r["name"],
                r["standort"],
                r["hoehe_u"],
                len(devs),
                round(ph["L1"] / 1000, 2),
                round(ph["L2"] / 1000, 2),
                round(ph["L3"] / 1000, 2),
                round(sum(ph.values()) / 1000, 2),
            ]
        )
        style_row(ws, ws.max_row)
    autowidth(ws)
    ws.freeze_panes = "A2"

    # Rack-Belegung
    ws = wb.create_sheet("Rack-Belegung")
    write_header(
        ws,
        [
            "Rack",
            "Standort",
            "U-von",
            "U-bis",
            "Hostname",
            "Typ",
            "IP",
            "Phase",
            "Anschlussleistung (W)",
            "Hersteller",
            "Modell",
            "Bemerkung",
        ],
    )
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            u_bis = (d.get("u_pos") or 0) + (d.get("u_h") or 1) - 1
            ws.append(
                [
                    r["name"],
                    r["standort"],
                    d.get("u_pos", "–"),
                    u_bis,
                    d["hostname"],
                    d["typ"],
                    d.get("ip", "–"),
                    d.get("phase", "–"),
                    d.get("watt") or "–",
                    d.get("hersteller", "–"),
                    d.get("modell", "–"),
                    d.get("bemerkung", ""),
                ]
            )
            style_row(
                ws,
                ws.max_row,
                fill_hex=DEVICE_COLOR.get(d.get("typ", "sonstige"), COL_ALT),
            )
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Ports & Interfaces
    ws = wb.create_sheet("Ports & Interfaces")
    write_header(
        ws,
        [
            "Rack",
            "Gerät",
            "Port",
            "Typ",
            "MAC",
            "Verbunden mit",
            "Ziel-Port",
            "Kabel-Nr",
            "Kabel-Typ",
            "m",
            "Farbe",
            "★",
        ],
    )
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            for iface in _ifaces_for_device(data, d["id"]):
                cable = _cable_for_iface(data, iface.get("kabel_id"))
                if cable:
                    is_von = cable.get("von_dev") == d["id"]
                    dst_id = cable.get("nach_dev") if is_von else cable.get("von_dev")
                    dst_dev = dm.get(dst_id)
                    dst_port = cable.get("nach_port" if is_von else "von_port", "–")
                    cross = dst_dev and dst_dev.get("rack_id") != r["id"]
                    dst_rn = rm.get(
                        dst_dev.get("rack_id") if dst_dev else None, {}
                    ).get("name", "")
                    xmark = f"★ {dst_rn}" if cross else ""
                    row = [
                        r["name"],
                        d["hostname"],
                        iface.get("port", "–"),
                        iface.get("typ", "–"),
                        iface.get("mac", "–"),
                        dst_dev["hostname"] if dst_dev else "–",
                        dst_port,
                        cable["nr"],
                        cable["typ"],
                        cable["laenge"],
                        cable.get("farbe", "–"),
                        xmark,
                    ]
                else:
                    row = [
                        r["name"],
                        d["hostname"],
                        iface.get("port", "–"),
                        iface.get("typ", "–"),
                        iface.get("mac", "–"),
                        "–",
                        "–",
                        "–",
                        "–",
                        "–",
                        "–",
                        "",
                    ]
                ws.append(row)
                style_row(ws, ws.max_row, fill_hex=COL_XRACK if row[-1] else COL_WHITE)
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Kabelliste
    ws = wb.create_sheet("Kabelliste")
    write_header(
        ws,
        [
            "Kabel-Nr",
            "Typ",
            "Länge m",
            "Farbe",
            "Von-Gerät",
            "Von-Rack",
            "Von-Port",
            "Nach-Gerät",
            "Nach-Rack",
            "Nach-Port",
            "Verlegt am",
            "Verlegt von",
            "★",
        ],
    )
    for c in sorted(data.get("cables", []), key=lambda x: x["nr"]):
        vd = dm.get(c.get("von_dev"))
        nd = dm.get(c.get("nach_dev"))
        vr = rm.get(vd.get("rack_id") if vd else None, {}).get("name", "–")
        nr = rm.get(nd.get("rack_id") if nd else None, {}).get("name", "–")
        cross = vd and nd and vd.get("rack_id") != nd.get("rack_id")
        ws.append(
            [
                c["nr"],
                c["typ"],
                c["laenge"],
                c.get("farbe", "–"),
                vd["hostname"] if vd else "–",
                vr,
                c.get("von_port", "–"),
                nd["hostname"] if nd else "–",
                nr,
                c.get("nach_port", "–"),
                c.get("verlegt_am", "–"),
                c.get("verlegt_von", "–"),
                "★ JA" if cross else "",
            ]
        )
        style_row(ws, ws.max_row, fill_hex=COL_XRACK if cross else COL_WHITE)
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # PDU-Belegung
    ws = wb.create_sheet("PDU-Belegung")
    write_header(
        ws,
        [
            "PDU",
            "Rack",
            "Steckdose",
            "Phase",
            "Steckdosentyp",
            "Max W",
            "Angeschlossenes Gerät",
            "Gerät-Rack",
            "Port",
            "Schaltbar",
            "Status",
        ],
    )
    for pdu in [d for d in data["devices"] if d.get("typ") == "pdu"]:
        rack_name = rm.get(pdu.get("rack_id"), {}).get("name", "–")
        outlets = _pdu_outlets(data, pdu["id"])
        if not outlets:
            ws.append(
                [
                    pdu["hostname"],
                    rack_name,
                    "–",
                    "–",
                    "–",
                    "–",
                    "–",
                    "–",
                    "–",
                    "–",
                    "kein Outlet",
                ]
            )
            style_row(ws, ws.max_row, fill_hex=COL_ALT)
            continue
        for o in outlets:
            dev = dm.get(o.get("connected_device_id"))
            dev_rack = rm.get(dev.get("rack_id") if dev else None, {}).get("name", "–")
            ws.append(
                [
                    pdu["hostname"],
                    rack_name,
                    o.get("outlet_name", "–"),
                    o.get("phase", "–"),
                    o.get("steckdosentyp", "–"),
                    o.get("max_watt", "–"),
                    dev["hostname"] if dev else "frei",
                    dev_rack,
                    o.get("connected_port", "–"),
                    "✓" if o.get("schaltbar") else "–",
                    "belegt" if dev else "frei",
                ]
            )
            fill = COL_WARN if not dev else PHASE_COL.get(o.get("phase", ""), COL_WHITE)
            style_row(ws, ws.max_row, fill_hex=fill)
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Import-Vorlage
    ws = wb.create_sheet("Import-Vorlage")
    ws["A1"] = "Hardware-Import-Vorlage"
    ws["A1"].font = Font(bold=True, size=11, name="Arial", color="AA0000")
    ws.append([])
    write_header(
        ws,
        [
            "hostname*",
            "typ*",
            "rack_name*",
            "u_position*",
            "u_hoehe",
            "ip_adresse",
            "phase",
            "tdp_watt",
            "hersteller",
            "modell",
            "seriennummer",
            "bemerkung",
        ],
        col_hex="2E4057",
    )
    ws.append(
        [
            "srv-beispiel-01",
            "server",
            "RACK-01",
            "40",
            "2",
            "192.168.1.100",
            "L1",
            "350",
            "Dell",
            "PowerEdge R650",
            "SN123456",
            "Beispielzeile",
        ]
    )
    style_row(ws, ws.max_row, fill_hex="FFFDE7")
    ws["A4"] = (
        "* Pflichtfelder  |  typ: server | switch | pdu | firewall | storage | sonstige  |  phase: L1 | L2 | L3"
    )
    ws["A4"].font = Font(size=8, name="Arial", color="888888", italic=True)
    autowidth(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _ods(data: dict) -> bytes:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = OpenDocumentSpreadsheet()
    dm = _dev_map(data)
    rm = _rack_map(data)

    def mkrow(values):
        tr = TableRow()
        for v in values:
            tc = TableCell()
            tc.addElement(P(text=str(v) if v is not None else ""))
            tr.addElement(tc)
        return tr

    def add_sheet(name, header, rows_fn):
        t = Table(name=name)
        t.addElement(mkrow(header))
        for row in rows_fn():
            t.addElement(mkrow(row))
        doc.spreadsheet.addElement(t)

    add_sheet(
        "Übersicht",
        ["Rack", "Standort", "HE", "Geräte", "L1 kW", "L2 kW", "L3 kW"],
        lambda: [
            [
                r["name"],
                r["standort"],
                r["hoehe_u"],
                len(_rack_devices(data, r["id"])),
                round(_phase_loads(data, r["id"])["L1"] / 1000, 2),
                round(_phase_loads(data, r["id"])["L2"] / 1000, 2),
                round(_phase_loads(data, r["id"])["L3"] / 1000, 2),
            ]
            for r in data["racks"]
        ],
    )

    def rack_rows():
        rows = []
        for r in data["racks"]:
            for d in _rack_devices(data, r["id"]):
                rows.append(
                    [
                        r["name"],
                        r["standort"],
                        d.get("u_pos", "–"),
                        (d.get("u_pos") or 0) + (d.get("u_h") or 1) - 1,
                        d["hostname"],
                        d["typ"],
                        d.get("ip", "–"),
                        d.get("phase", "–"),
                        d.get("watt", "–"),
                        d.get("hersteller", "–"),
                        d.get("modell", "–"),
                    ]
                )
        return rows

    add_sheet(
        "Rack-Belegung",
        [
            "Rack",
            "Standort",
            "U-von",
            "U-bis",
            "Hostname",
            "Typ",
            "IP",
            "Phase",
            "Anschlussleistung (W)",
            "Hersteller",
            "Modell",
        ],
        rack_rows,
    )

    def iface_rows():
        rows = []
        for r in data["racks"]:
            for d in _rack_devices(data, r["id"]):
                for iface in _ifaces_for_device(data, d["id"]):
                    cable = _cable_for_iface(data, iface.get("kabel_id"))
                    if cable:
                        is_von = cable.get("von_dev") == d["id"]
                        dst_id = (
                            cable.get("nach_dev") if is_von else cable.get("von_dev")
                        )
                        dst_dev = dm.get(dst_id)
                        dst_port = cable.get("nach_port" if is_von else "von_port", "–")
                        cross = (
                            "★"
                            if (dst_dev and dst_dev.get("rack_id") != r["id"])
                            else ""
                        )
                        rows.append(
                            [
                                r["name"],
                                d["hostname"],
                                iface.get("port", "–"),
                                iface.get("typ", "–"),
                                iface.get("mac", "–"),
                                dst_dev["hostname"] if dst_dev else "–",
                                dst_port,
                                cable["nr"],
                                cable["typ"],
                                cable["laenge"],
                                cable.get("farbe", "–"),
                                cross,
                            ]
                        )
                    else:
                        rows.append(
                            [
                                r["name"],
                                d["hostname"],
                                iface.get("port", "–"),
                                iface.get("typ", "–"),
                                iface.get("mac", "–"),
                                "–",
                                "–",
                                "–",
                                "–",
                                "–",
                                "–",
                                "",
                            ]
                        )
        return rows

    add_sheet(
        "Ports & Interfaces",
        [
            "Rack",
            "Gerät",
            "Port",
            "Typ",
            "MAC",
            "Verbunden mit",
            "Ziel-Port",
            "Kabel-Nr",
            "Kabel-Typ",
            "m",
            "Farbe",
            "★",
        ],
        iface_rows,
    )

    def cable_rows():
        rows = []
        for c in sorted(data.get("cables", []), key=lambda x: x["nr"]):
            vd = dm.get(c.get("von_dev"))
            nd = dm.get(c.get("nach_dev"))
            vr = rm.get(vd.get("rack_id") if vd else None, {}).get("name", "–")
            nr = rm.get(nd.get("rack_id") if nd else None, {}).get("name", "–")
            cross = (
                "★" if (vd and nd and vd.get("rack_id") != nd.get("rack_id")) else ""
            )
            rows.append(
                [
                    c["nr"],
                    c["typ"],
                    c["laenge"],
                    c.get("farbe", "–"),
                    vd["hostname"] if vd else "–",
                    vr,
                    c.get("von_port", "–"),
                    nd["hostname"] if nd else "–",
                    nr,
                    c.get("nach_port", "–"),
                    cross,
                ]
            )
        return rows

    add_sheet(
        "Kabelliste",
        [
            "Kabel-Nr",
            "Typ",
            "m",
            "Farbe",
            "Von-Gerät",
            "Von-Rack",
            "Von-Port",
            "Nach-Gerät",
            "Nach-Rack",
            "Nach-Port",
            "★",
        ],
        cable_rows,
    )

    def pdu_rows():
        rows = []
        for pdu in [d for d in data["devices"] if d.get("typ") == "pdu"]:
            rn = rm.get(pdu.get("rack_id"), {}).get("name", "–")
            for o in _pdu_outlets(data, pdu["id"]):
                dev = dm.get(o.get("connected_device_id"))
                rows.append(
                    [
                        pdu["hostname"],
                        rn,
                        o.get("outlet_name", "–"),
                        o.get("phase", "–"),
                        o.get("steckdosentyp", "–"),
                        o.get("max_watt", "–"),
                        dev["hostname"] if dev else "frei",
                        o.get("connected_port", "–"),
                        "JA" if o.get("schaltbar") else "NEIN",
                    ]
                )
        return rows

    add_sheet(
        "PDU-Belegung",
        [
            "PDU",
            "Rack",
            "Steckdose",
            "Phase",
            "Typ",
            "Max W",
            "Gerät",
            "Port",
            "Schaltbar",
        ],
        pdu_rows,
    )

    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _csv_zip(data: dict) -> bytes:
    dm = _dev_map(data)
    rm = _rack_map(data)

    def to_csv(header, rows):
        buf = io.StringIO()
        w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        w.writerow(header)
        w.writerows(rows)
        return buf.getvalue().encode("utf-8-sig")

    ov = [
        [
            r["name"],
            r["standort"],
            r["hoehe_u"],
            len(_rack_devices(data, r["id"])),
            round(_phase_loads(data, r["id"])["L1"] / 1000, 2),
            round(_phase_loads(data, r["id"])["L2"] / 1000, 2),
            round(_phase_loads(data, r["id"])["L3"] / 1000, 2),
        ]
        for r in data["racks"]
    ]

    dev_rows = []
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            dev_rows.append(
                [
                    r["name"],
                    d.get("u_pos", ""),
                    d.get("u_h", 1),
                    d["hostname"],
                    d["typ"],
                    d.get("ip", ""),
                    d.get("phase", ""),
                    d.get("watt", ""),
                    d.get("hersteller", ""),
                    d.get("modell", ""),
                ]
            )

    port_rows = []
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            for iface in _ifaces_for_device(data, d["id"]):
                cable = _cable_for_iface(data, iface.get("kabel_id"))
                if cable:
                    is_von = cable.get("von_dev") == d["id"]
                    dst_id = cable.get("nach_dev") if is_von else cable.get("von_dev")
                    dst_dev = dm.get(dst_id)
                    dst_port = cable.get("nach_port" if is_von else "von_port", "")
                    cross = (
                        "JA" if (dst_dev and dst_dev.get("rack_id") != r["id"]) else ""
                    )
                    port_rows.append(
                        [
                            r["name"],
                            d["hostname"],
                            iface.get("port", ""),
                            iface.get("typ", ""),
                            iface.get("mac", ""),
                            dst_dev["hostname"] if dst_dev else "",
                            dst_port,
                            cable["nr"],
                            cable["typ"],
                            cable["laenge"],
                            cable.get("farbe", ""),
                            cross,
                        ]
                    )
                else:
                    port_rows.append(
                        [
                            r["name"],
                            d["hostname"],
                            iface.get("port", ""),
                            iface.get("typ", ""),
                            iface.get("mac", ""),
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ]
                    )

    cable_rows = []
    for c in sorted(data.get("cables", []), key=lambda x: x["nr"]):
        vd = dm.get(c.get("von_dev"))
        nd = dm.get(c.get("nach_dev"))
        vr = rm.get(vd.get("rack_id") if vd else None, {}).get("name", "")
        nr = rm.get(nd.get("rack_id") if nd else None, {}).get("name", "")
        cross = "JA" if (vd and nd and vd.get("rack_id") != nd.get("rack_id")) else ""
        cable_rows.append(
            [
                c["nr"],
                c["typ"],
                c["laenge"],
                c.get("farbe", ""),
                vd["hostname"] if vd else "",
                vr,
                c.get("von_port", ""),
                nd["hostname"] if nd else "",
                nr,
                c.get("nach_port", ""),
                c.get("verlegt_am", ""),
                c.get("verlegt_von", ""),
                cross,
            ]
        )

    pdu_rows = []
    for pdu in [d for d in data["devices"] if d.get("typ") == "pdu"]:
        rn = rm.get(pdu.get("rack_id"), {}).get("name", "")
        for o in _pdu_outlets(data, pdu["id"]):
            dev = dm.get(o.get("connected_device_id"))
            pdu_rows.append(
                [
                    pdu["hostname"],
                    rn,
                    o.get("outlet_name", ""),
                    o.get("phase", ""),
                    o.get("steckdosentyp", ""),
                    o.get("max_watt", ""),
                    dev["hostname"] if dev else "frei",
                    o.get("connected_port", ""),
                    "JA" if o.get("schaltbar") else "NEIN",
                ]
            )

    ts = datetime.now().strftime("%Y%m%d_%H%M")
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        zf.writestr(
            f"kaitix_{ts}_uebersicht.csv",
            to_csv(
                ["rack", "standort", "he", "geraete", "l1_kw", "l2_kw", "l3_kw"], ov
            ),
        )
        zf.writestr(
            f"kaitix_{ts}_geraete.csv",
            to_csv(
                [
                    "rack",
                    "u_position",
                    "u_hoehe",
                    "hostname",
                    "typ",
                    "ip",
                    "phase",
                    "anschlussleistung_w",
                    "hersteller",
                    "modell",
                ],
                dev_rows,
            ),
        )
        zf.writestr(
            f"kaitix_{ts}_ports.csv",
            to_csv(
                [
                    "rack",
                    "geraet",
                    "port",
                    "typ",
                    "mac",
                    "ziel_geraet",
                    "ziel_port",
                    "kabel_nr",
                    "kabel_typ",
                    "laenge_m",
                    "farbe",
                    "rack_uebergreifend",
                ],
                port_rows,
            ),
        )
        zf.writestr(
            f"kaitix_{ts}_kabel.csv",
            to_csv(
                [
                    "kabel_nr",
                    "typ",
                    "laenge_m",
                    "farbe",
                    "von_geraet",
                    "von_rack",
                    "von_port",
                    "nach_geraet",
                    "nach_rack",
                    "nach_port",
                    "verlegt_am",
                    "verlegt_von",
                    "rack_uebergreifend",
                ],
                cable_rows,
            ),
        )
        zf.writestr(
            f"kaitix_{ts}_pdu.csv",
            to_csv(
                [
                    "pdu",
                    "rack",
                    "steckdose",
                    "phase",
                    "steckdosentyp",
                    "max_watt",
                    "geraet",
                    "port",
                    "schaltbar",
                ],
                pdu_rows,
            ),
        )
    return zip_buf.getvalue()


Format = Literal["xlsx", "ods", "csv"]


def build_export(data: dict, fmt: Format) -> tuple[bytes, str, str]:
    """
    Gibt zurück: (bytes, mimetype, filename)
    data muss enthalten: racks, devices, interfaces, cables, pdu_outlets
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    if fmt == "xlsx":
        return (
            _xlsx(data),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            f"kaitix_dokumentation_{ts}.xlsx",
        )
    elif fmt == "ods":
        return (
            _ods(data),
            "application/vnd.oasis.opendocument.spreadsheet",
            f"kaitix_dokumentation_{ts}.ods",
        )
    else:
        return (_csv_zip(data), "application/zip", f"kaitix_dokumentation_{ts}.zip")


# ---------------------------------------------------------------------------
# Einzel-Exporte (Rack-Inventar, Ports & Interfaces, PDU-Belegung, Kabelliste)
# ---------------------------------------------------------------------------


def _rack_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)

    def hf():
        return Font(bold=True, color="FFFFFF", name="Arial", size=10)

    def hfill(c):
        return PatternFill("solid", fgColor=c)

    def cfill(c):
        return PatternFill("solid", fgColor=c)

    def aln(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=True)

    def write_header(ws, cols, col_hex=COL_HEADER):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = hf()
            cell.fill = hfill(col_hex)
            cell.border = brd
            cell.alignment = aln("center")
        ws.row_dimensions[1].height = 20

    def style_row(ws, row_idx, fill_hex=None, bold=False):
        fill = cfill(fill_hex) if fill_hex else None
        for cell in ws[row_idx]:
            cell.font = Font(bold=bold, name="Arial", size=9)
            cell.border = brd
            cell.alignment = aln()
            if fill:
                cell.fill = fill

    def autowidth(ws, mn=8, mx=40):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(w + 2, mn), mx
            )

    ws = wb.create_sheet("Rack-Inventar")
    write_header(
        ws,
        [
            "Rack",
            "Standort",
            "U-von",
            "U-bis",
            "Hostname",
            "Typ",
            "IP",
            "Phase",
            "Anschlussleistung (W)",
            "Hersteller",
            "Modell",
            "Bemerkung",
        ],
    )
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            u_bis = (d.get("u_pos") or 0) + (d.get("u_h") or 1) - 1
            ws.append(
                [
                    r["name"],
                    r["standort"],
                    d.get("u_pos", "–"),
                    u_bis,
                    d["hostname"],
                    d["typ"],
                    d.get("ip", "–"),
                    d.get("phase", "–"),
                    d.get("watt") or "–",
                    d.get("hersteller", "–"),
                    d.get("modell", "–"),
                    d.get("bemerkung", ""),
                ]
            )
            style_row(
                ws,
                ws.max_row,
                fill_hex=DEVICE_COLOR.get(d.get("typ", "sonstige"), COL_ALT),
            )
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _rack_ods(data: dict) -> bytes:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = OpenDocumentSpreadsheet()

    def mkrow(values):
        tr = TableRow()
        for v in values:
            tc = TableCell()
            tc.addElement(P(text=str(v) if v is not None else ""))
            tr.addElement(tc)
        return tr

    t = Table(name="Rack-Inventar")
    t.addElement(
        mkrow(
            [
                "Rack",
                "Standort",
                "U-von",
                "U-bis",
                "Hostname",
                "Typ",
                "IP",
                "Phase",
                "Anschlussleistung (W)",
                "Hersteller",
                "Modell",
            ]
        )
    )
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            t.addElement(
                mkrow(
                    [
                        r["name"],
                        r["standort"],
                        d.get("u_pos", "–"),
                        (d.get("u_pos") or 0) + (d.get("u_h") or 1) - 1,
                        d["hostname"],
                        d["typ"],
                        d.get("ip", "–"),
                        d.get("phase", "–"),
                        d.get("watt", "–"),
                        d.get("hersteller", "–"),
                        d.get("modell", "–"),
                    ]
                )
            )
    doc.spreadsheet.addElement(t)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _rack_csv(data: dict) -> bytes:
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(
        [
            "rack",
            "standort",
            "u_von",
            "u_bis",
            "hostname",
            "typ",
            "ip",
            "phase",
            "anschlussleistung_w",
            "hersteller",
            "modell",
            "bemerkung",
        ]
    )
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            w.writerow(
                [
                    r["name"],
                    r["standort"],
                    d.get("u_pos", ""),
                    (d.get("u_pos") or 0) + (d.get("u_h") or 1) - 1,
                    d["hostname"],
                    d["typ"],
                    d.get("ip", ""),
                    d.get("phase", ""),
                    d.get("watt", ""),
                    d.get("hersteller", ""),
                    d.get("modell", ""),
                    d.get("bemerkung", ""),
                ]
            )
    return buf.getvalue().encode("utf-8-sig")


def _iface_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    dm = _dev_map(data)
    rm = _rack_map(data)

    def hf():
        return Font(bold=True, color="FFFFFF", name="Arial", size=10)

    def hfill(c):
        return PatternFill("solid", fgColor=c)

    def cfill(c):
        return PatternFill("solid", fgColor=c)

    def aln(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=True)

    def write_header(ws, cols, col_hex=COL_HEADER):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = hf()
            cell.fill = hfill(col_hex)
            cell.border = brd
            cell.alignment = aln("center")
        ws.row_dimensions[1].height = 20

    def style_row(ws, row_idx, fill_hex=None, bold=False):
        fill = cfill(fill_hex) if fill_hex else None
        for cell in ws[row_idx]:
            cell.font = Font(bold=bold, name="Arial", size=9)
            cell.border = brd
            cell.alignment = aln()
            if fill:
                cell.fill = fill

    def autowidth(ws, mn=8, mx=40):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(w + 2, mn), mx
            )

    ws = wb.create_sheet("Ports & Interfaces")
    write_header(
        ws,
        [
            "Rack",
            "Gerät",
            "Port",
            "Typ",
            "MAC",
            "Verbunden mit",
            "Ziel-Port",
            "Kabel-Nr",
            "Kabel-Typ",
            "m",
            "Farbe",
            "★",
        ],
    )
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            for iface in _ifaces_for_device(data, d["id"]):
                cable = _cable_for_iface(data, iface.get("kabel_id"))
                if cable:
                    is_von = cable.get("von_dev") == d["id"]
                    dst_id = cable.get("nach_dev") if is_von else cable.get("von_dev")
                    dst_dev = dm.get(dst_id)
                    dst_port = cable.get("nach_port" if is_von else "von_port", "–")
                    cross = dst_dev and dst_dev.get("rack_id") != r["id"]
                    dst_rn = rm.get(
                        dst_dev.get("rack_id") if dst_dev else None, {}
                    ).get("name", "")
                    xmark = f"★ {dst_rn}" if cross else ""
                    row = [
                        r["name"],
                        d["hostname"],
                        iface.get("port", "–"),
                        iface.get("typ", "–"),
                        iface.get("mac", "–"),
                        dst_dev["hostname"] if dst_dev else "–",
                        dst_port,
                        cable["nr"],
                        cable["typ"],
                        cable["laenge"],
                        cable.get("farbe", "–"),
                        xmark,
                    ]
                else:
                    row = [
                        r["name"],
                        d["hostname"],
                        iface.get("port", "–"),
                        iface.get("typ", "–"),
                        iface.get("mac", "–"),
                        "–",
                        "–",
                        "–",
                        "–",
                        "–",
                        "–",
                        "",
                    ]
                ws.append(row)
                style_row(ws, ws.max_row, fill_hex=COL_XRACK if row[-1] else COL_WHITE)
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _iface_ods(data: dict) -> bytes:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = OpenDocumentSpreadsheet()
    dm = _dev_map(data)

    def mkrow(values):
        tr = TableRow()
        for v in values:
            tc = TableCell()
            tc.addElement(P(text=str(v) if v is not None else ""))
            tr.addElement(tc)
        return tr

    t = Table(name="Ports & Interfaces")
    t.addElement(
        mkrow(
            [
                "Rack",
                "Gerät",
                "Port",
                "Typ",
                "MAC",
                "Verbunden mit",
                "Ziel-Port",
                "Kabel-Nr",
                "Kabel-Typ",
                "m",
                "Farbe",
                "★",
            ]
        )
    )
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            for iface in _ifaces_for_device(data, d["id"]):
                cable = _cable_for_iface(data, iface.get("kabel_id"))
                if cable:
                    is_von = cable.get("von_dev") == d["id"]
                    dst_id = cable.get("nach_dev") if is_von else cable.get("von_dev")
                    dst_dev = dm.get(dst_id)
                    dst_port = cable.get("nach_port" if is_von else "von_port", "–")
                    cross = (
                        "★" if (dst_dev and dst_dev.get("rack_id") != r["id"]) else ""
                    )
                    t.addElement(
                        mkrow(
                            [
                                r["name"],
                                d["hostname"],
                                iface.get("port", "–"),
                                iface.get("typ", "–"),
                                iface.get("mac", "–"),
                                dst_dev["hostname"] if dst_dev else "–",
                                dst_port,
                                cable["nr"],
                                cable["typ"],
                                cable["laenge"],
                                cable.get("farbe", "–"),
                                cross,
                            ]
                        )
                    )
                else:
                    t.addElement(
                        mkrow(
                            [
                                r["name"],
                                d["hostname"],
                                iface.get("port", "–"),
                                iface.get("typ", "–"),
                                iface.get("mac", "–"),
                                "–",
                                "–",
                                "–",
                                "–",
                                "–",
                                "–",
                                "",
                            ]
                        )
                    )
    doc.spreadsheet.addElement(t)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _iface_csv(data: dict) -> bytes:
    dm = _dev_map(data)
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(
        [
            "rack",
            "geraet",
            "port",
            "typ",
            "mac",
            "ziel_geraet",
            "ziel_port",
            "kabel_nr",
            "kabel_typ",
            "laenge_m",
            "farbe",
            "rack_uebergreifend",
        ]
    )
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            for iface in _ifaces_for_device(data, d["id"]):
                cable = _cable_for_iface(data, iface.get("kabel_id"))
                if cable:
                    is_von = cable.get("von_dev") == d["id"]
                    dst_id = cable.get("nach_dev") if is_von else cable.get("von_dev")
                    dst_dev = dm.get(dst_id)
                    dst_port = cable.get("nach_port" if is_von else "von_port", "")
                    cross = (
                        "JA" if (dst_dev and dst_dev.get("rack_id") != r["id"]) else ""
                    )
                    w.writerow(
                        [
                            r["name"],
                            d["hostname"],
                            iface.get("port", ""),
                            iface.get("typ", ""),
                            iface.get("mac", ""),
                            dst_dev["hostname"] if dst_dev else "",
                            dst_port,
                            cable["nr"],
                            cable["typ"],
                            cable["laenge"],
                            cable.get("farbe", ""),
                            cross,
                        ]
                    )
                else:
                    w.writerow(
                        [
                            r["name"],
                            d["hostname"],
                            iface.get("port", ""),
                            iface.get("typ", ""),
                            iface.get("mac", ""),
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                            "",
                        ]
                    )
    return buf.getvalue().encode("utf-8-sig")


def _pdu_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    dm = _dev_map(data)
    rm = _rack_map(data)

    def hf():
        return Font(bold=True, color="FFFFFF", name="Arial", size=10)

    def hfill(c):
        return PatternFill("solid", fgColor=c)

    def cfill(c):
        return PatternFill("solid", fgColor=c)

    def aln(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=True)

    def write_header(ws, cols, col_hex=COL_HEADER):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = hf()
            cell.fill = hfill(col_hex)
            cell.border = brd
            cell.alignment = aln("center")
        ws.row_dimensions[1].height = 20

    def style_row(ws, row_idx, fill_hex=None, bold=False):
        fill = cfill(fill_hex) if fill_hex else None
        for cell in ws[row_idx]:
            cell.font = Font(bold=bold, name="Arial", size=9)
            cell.border = brd
            cell.alignment = aln()
            if fill:
                cell.fill = fill

    def autowidth(ws, mn=8, mx=40):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(w + 2, mn), mx
            )

    ws = wb.create_sheet("PDU-Belegung")
    write_header(
        ws,
        [
            "PDU",
            "Rack",
            "Steckdose",
            "Phase",
            "Steckdosentyp",
            "Max W",
            "Angeschlossenes Gerät",
            "Gerät-Rack",
            "Port",
            "Schaltbar",
            "Status",
        ],
    )
    for pdu in [d for d in data["devices"] if d.get("typ") == "pdu"]:
        rack_name = rm.get(pdu.get("rack_id"), {}).get("name", "–")
        outlets = _pdu_outlets(data, pdu["id"])
        if not outlets:
            ws.append(
                [
                    pdu["hostname"],
                    rack_name,
                    "–",
                    "–",
                    "–",
                    "–",
                    "–",
                    "–",
                    "–",
                    "–",
                    "kein Outlet",
                ]
            )
            style_row(ws, ws.max_row, fill_hex=COL_ALT)
            continue
        for o in outlets:
            dev = dm.get(o.get("connected_device_id"))
            dev_rack = rm.get(dev.get("rack_id") if dev else None, {}).get("name", "–")
            ws.append(
                [
                    pdu["hostname"],
                    rack_name,
                    o.get("outlet_name", "–"),
                    o.get("phase", "–"),
                    o.get("steckdosentyp", "–"),
                    o.get("max_watt", "–"),
                    dev["hostname"] if dev else "frei",
                    dev_rack,
                    o.get("connected_port", "–"),
                    "✓" if o.get("schaltbar") else "–",
                    "belegt" if dev else "frei",
                ]
            )
            fill = COL_WARN if not dev else PHASE_COL.get(o.get("phase", ""), COL_WHITE)
            style_row(ws, ws.max_row, fill_hex=fill)
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdu_ods(data: dict) -> bytes:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = OpenDocumentSpreadsheet()
    dm = _dev_map(data)
    rm = _rack_map(data)

    def mkrow(values):
        tr = TableRow()
        for v in values:
            tc = TableCell()
            tc.addElement(P(text=str(v) if v is not None else ""))
            tr.addElement(tc)
        return tr

    t = Table(name="PDU-Belegung")
    t.addElement(
        mkrow(
            [
                "PDU",
                "Rack",
                "Steckdose",
                "Phase",
                "Typ",
                "Max W",
                "Gerät",
                "Gerät-Rack",
                "Port",
                "Schaltbar",
                "Status",
            ]
        )
    )
    for pdu in [d for d in data["devices"] if d.get("typ") == "pdu"]:
        rn = rm.get(pdu.get("rack_id"), {}).get("name", "–")
        for o in _pdu_outlets(data, pdu["id"]):
            dev = dm.get(o.get("connected_device_id"))
            t.addElement(
                mkrow(
                    [
                        pdu["hostname"],
                        rn,
                        o.get("outlet_name", "–"),
                        o.get("phase", "–"),
                        o.get("steckdosentyp", "–"),
                        o.get("max_watt", "–"),
                        dev["hostname"] if dev else "frei",
                        rm.get(dev.get("rack_id") if dev else None, {}).get(
                            "name", "–"
                        ),
                        o.get("connected_port", "–"),
                        "JA" if o.get("schaltbar") else "NEIN",
                        "belegt" if dev else "frei",
                    ]
                )
            )
    doc.spreadsheet.addElement(t)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _pdu_csv(data: dict) -> bytes:
    dm = _dev_map(data)
    rm = _rack_map(data)
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(
        [
            "pdu",
            "rack",
            "steckdose",
            "phase",
            "steckdosentyp",
            "max_watt",
            "geraet",
            "geraet_rack",
            "port",
            "schaltbar",
            "status",
        ]
    )
    for pdu in [d for d in data["devices"] if d.get("typ") == "pdu"]:
        rn = rm.get(pdu.get("rack_id"), {}).get("name", "")
        for o in _pdu_outlets(data, pdu["id"]):
            dev = dm.get(o.get("connected_device_id"))
            w.writerow(
                [
                    pdu["hostname"],
                    rn,
                    o.get("outlet_name", ""),
                    o.get("phase", ""),
                    o.get("steckdosentyp", ""),
                    o.get("max_watt", ""),
                    dev["hostname"] if dev else "frei",
                    rm.get(dev.get("rack_id") if dev else None, {}).get("name", ""),
                    o.get("connected_port", ""),
                    "JA" if o.get("schaltbar") else "NEIN",
                    "belegt" if dev else "frei",
                ]
            )
    return buf.getvalue().encode("utf-8-sig")


def _cable_xlsx(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    dm = _dev_map(data)
    rm = _rack_map(data)

    def hf():
        return Font(bold=True, color="FFFFFF", name="Arial", size=10)

    def hfill(c):
        return PatternFill("solid", fgColor=c)

    def cfill(c):
        return PatternFill("solid", fgColor=c)

    def aln(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=True)

    def write_header(ws, cols, col_hex=COL_HEADER):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = hf()
            cell.fill = hfill(col_hex)
            cell.border = brd
            cell.alignment = aln("center")
        ws.row_dimensions[1].height = 20

    def style_row(ws, row_idx, fill_hex=None, bold=False):
        fill = cfill(fill_hex) if fill_hex else None
        for cell in ws[row_idx]:
            cell.font = Font(bold=bold, name="Arial", size=9)
            cell.border = brd
            cell.alignment = aln()
            if fill:
                cell.fill = fill

    def autowidth(ws, mn=8, mx=40):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(w + 2, mn), mx
            )

    ws = wb.create_sheet("Kabelliste")
    write_header(
        ws,
        [
            "Kabel-Nr",
            "Typ",
            "Länge m",
            "Farbe",
            "Von-Gerät",
            "Von-Rack",
            "Von-Port",
            "Nach-Gerät",
            "Nach-Rack",
            "Nach-Port",
            "Verlegt am",
            "Verlegt von",
            "★",
        ],
    )
    for c in sorted(data.get("cables", []), key=lambda x: x["nr"]):
        vd = dm.get(c.get("von_dev"))
        nd = dm.get(c.get("nach_dev"))
        vr = rm.get(vd.get("rack_id") if vd else None, {}).get("name", "–")
        nr = rm.get(nd.get("rack_id") if nd else None, {}).get("name", "–")
        cross = vd and nd and vd.get("rack_id") != nd.get("rack_id")
        ws.append(
            [
                c["nr"],
                c["typ"],
                c["laenge"],
                c.get("farbe", "–"),
                vd["hostname"] if vd else "–",
                vr,
                c.get("von_port", "–"),
                nd["hostname"] if nd else "–",
                nr,
                c.get("nach_port", "–"),
                c.get("verlegt_am", "–"),
                c.get("verlegt_von", "–"),
                "★ JA" if cross else "",
            ]
        )
        style_row(ws, ws.max_row, fill_hex=COL_XRACK if cross else COL_WHITE)
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cable_ods(data: dict) -> bytes:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P

    doc = OpenDocumentSpreadsheet()
    dm = _dev_map(data)
    rm = _rack_map(data)

    def mkrow(values):
        tr = TableRow()
        for v in values:
            tc = TableCell()
            tc.addElement(P(text=str(v) if v is not None else ""))
            tr.addElement(tc)
        return tr

    t = Table(name="Kabelliste")
    t.addElement(
        mkrow(
            [
                "Kabel-Nr",
                "Typ",
                "m",
                "Farbe",
                "Von-Gerät",
                "Von-Rack",
                "Von-Port",
                "Nach-Gerät",
                "Nach-Rack",
                "Nach-Port",
                "★",
            ]
        )
    )
    for c in sorted(data.get("cables", []), key=lambda x: x["nr"]):
        vd = dm.get(c.get("von_dev"))
        nd = dm.get(c.get("nach_dev"))
        vr = rm.get(vd.get("rack_id") if vd else None, {}).get("name", "–")
        nr = rm.get(nd.get("rack_id") if nd else None, {}).get("name", "–")
        cross = "★" if (vd and nd and vd.get("rack_id") != nd.get("rack_id")) else ""
        t.addElement(
            mkrow(
                [
                    c["nr"],
                    c["typ"],
                    c["laenge"],
                    c.get("farbe", "–"),
                    vd["hostname"] if vd else "–",
                    vr,
                    c.get("von_port", "–"),
                    nd["hostname"] if nd else "–",
                    nr,
                    c.get("nach_port", "–"),
                    cross,
                ]
            )
        )
    doc.spreadsheet.addElement(t)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _cable_csv(data: dict) -> bytes:
    dm = _dev_map(data)
    rm = _rack_map(data)
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(
        [
            "kabel_nr",
            "typ",
            "laenge_m",
            "farbe",
            "von_geraet",
            "von_rack",
            "von_port",
            "nach_geraet",
            "nach_rack",
            "nach_port",
            "verlegt_am",
            "verlegt_von",
            "rack_uebergreifend",
        ]
    )
    for c in sorted(data.get("cables", []), key=lambda x: x["nr"]):
        vd = dm.get(c.get("von_dev"))
        nd = dm.get(c.get("nach_dev"))
        vr = rm.get(vd.get("rack_id") if vd else None, {}).get("name", "")
        nr = rm.get(nd.get("rack_id") if nd else None, {}).get("name", "")
        cross = "JA" if (vd and nd and vd.get("rack_id") != nd.get("rack_id")) else ""
        w.writerow(
            [
                c["nr"],
                c["typ"],
                c["laenge"],
                c.get("farbe", ""),
                vd["hostname"] if vd else "",
                vr,
                c.get("von_port", ""),
                nd["hostname"] if nd else "",
                nr,
                c.get("nach_port", ""),
                c.get("verlegt_am", ""),
                c.get("verlegt_von", ""),
                cross,
            ]
        )
    return buf.getvalue().encode("utf-8-sig")


Sheet = Literal["racks", "interfaces", "pdus", "cables"]


def build_single_export(
    data: dict, sheet: Sheet, fmt: Format
) -> tuple[bytes, str, str]:
    """
    Einzel-Export für ein bestimmtes Sheet.
    Gibt zurück: (bytes, mimetype, filename)
    """
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    builders = {
        "racks": {
            "xlsx": (
                _rack_xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"kaitix_rack_inventar_{ts}.xlsx",
            ),
            "ods": (
                _rack_ods,
                "application/vnd.oasis.opendocument.spreadsheet",
                f"kaitix_rack_inventar_{ts}.ods",
            ),
            "csv": (_rack_csv, "text/csv", f"kaitix_rack_inventar_{ts}.csv"),
        },
        "interfaces": {
            "xlsx": (
                _iface_xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"kaitix_ports_interfaces_{ts}.xlsx",
            ),
            "ods": (
                _iface_ods,
                "application/vnd.oasis.opendocument.spreadsheet",
                f"kaitix_ports_interfaces_{ts}.ods",
            ),
            "csv": (_iface_csv, "text/csv", f"kaitix_ports_interfaces_{ts}.csv"),
        },
        "pdus": {
            "xlsx": (
                _pdu_xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"kaitix_pdu_belegung_{ts}.xlsx",
            ),
            "ods": (
                _pdu_ods,
                "application/vnd.oasis.opendocument.spreadsheet",
                f"kaitix_pdu_belegung_{ts}.ods",
            ),
            "csv": (_pdu_csv, "text/csv", f"kaitix_pdu_belegung_{ts}.csv"),
        },
        "cables": {
            "xlsx": (
                _cable_xlsx,
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                f"kaitix_kabelliste_{ts}.xlsx",
            ),
            "ods": (
                _cable_ods,
                "application/vnd.oasis.opendocument.spreadsheet",
                f"kaitix_kabelliste_{ts}.ods",
            ),
            "csv": (_cable_csv, "text/csv", f"kaitix_kabelliste_{ts}.csv"),
        },
    }
    builder_fn, mime, fname = builders[sheet][fmt]
    return (builder_fn(data), mime, fname)
