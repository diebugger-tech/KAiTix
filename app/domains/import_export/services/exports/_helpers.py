import csv
import io
from datetime import datetime
from typing import Literal

COL_HEADER = "1A1A2E"
COL_SERVER = "D6E4F7"
COL_SWITCH = "D6F0E8"
COL_PDU = "FFF3CD"
COL_FIREWALL = "FFE0D6"
COL_STORAGE = "EDE0F7"
COL_XRACK = "FFD6C0"
COL_WARN = "FFC7C7"
COL_WHITE = "FFFFFF"
COL_ALT = "F5F5F5"

DEVICE_COLOR = {
    "server": COL_SERVER,
    "switch": COL_SWITCH,
    "pdu": COL_PDU,
    "firewall": COL_FIREWALL,
    "storage": COL_STORAGE,
    "sonstige": COL_ALT,
}
PHASE_COL = {"L1": "D6EAF8", "L2": "D5F5E3", "L3": "FDEBD0"}

Format = Literal["xlsx", "ods", "csv"]
Sheet = Literal["racks", "interfaces", "pdus", "cables"]


def _dev_map(data):
    return {d["id"]: d for d in data["devices"]}


def _rack_map(data):
    return {r["id"]: r for r in data["racks"]}


def _phase_loads(data, rack_id):
    ph = {"L1": 0.0, "L2": 0.0, "L3": 0.0}
    for d in data["devices"]:
        if d.get("rack_id") == rack_id and d.get("phase") in ph:
            ph[d["phase"]] += d.get("watt", 0)
    return ph


def _rack_devices(data, rack_id):
    return sorted(
        [d for d in data["devices"] if d.get("rack_id") == rack_id],
        key=lambda d: -(d.get("u_pos") or 0),
    )


def _ifaces_for_device(data, dev_id):
    return [i for i in data.get("interfaces", []) if i.get("dev_id") == dev_id]


def _cable_for_iface(data, kabel_id):
    if not kabel_id:
        return None
    return next((c for c in data.get("cables", []) if c["id"] == kabel_id), None)


def _pdu_outlets(data, pdu_id):
    return [o for o in data.get("pdu_outlets", []) if o.get("pdu_id") == pdu_id]


def _ts():
    return datetime.now().strftime("%Y%m%d_%H%M")


def _to_csv(header, rows):
    buf = io.StringIO()
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(header)
    w.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")


def _make_xlsx_helpers(brd):
    from openpyxl.styles import Font, PatternFill, Alignment  # type: ignore
    from openpyxl.utils import get_column_letter  # type: ignore

    def hf():
        return Font(bold=True, color="FFFFFF", name="Arial", size=10)

    def hfill(c):
        return PatternFill("solid", fgColor=c)

    def cfill(c):
        return PatternFill("solid", fgColor=c)

    def aln(h="left"):
        return Alignment(horizontal=h, vertical="center", wrap_text=True)

    def write_header(ws, cols, col_hex=COL_HEADER):
        ws.append(cols)
        for cell in ws[1]:
            cell.font = hf()
            cell.fill = hfill(col_hex)
            cell.border = brd
            cell.alignment = aln("center")
        ws.row_dimensions[1].height = 20

    def style_row(ws, row_idx, fill_hex=None, bold=False):
        fill = cfill(fill_hex) if fill_hex else None
        for cell in ws[row_idx]:
            cell.font = Font(bold=bold, name="Arial", size=9)
            cell.border = brd
            cell.alignment = aln()
            if fill:
                cell.fill = fill

    def autowidth(ws, mn=8, mx=40):
        for col in ws.columns:
            w = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(
                max(w + 2, mn), mx
            )

    return write_header, style_row, autowidth


def _make_ods_mkrow():
    from odf.table import TableRow, TableCell  # type: ignore
    from odf.text import P  # type: ignore

    def mkrow(values):
        tr = TableRow()
        for v in values:
            tc = TableCell()
            tc.addElement(P(text=str(v) if v is not None else ""))
            tr.addElement(tc)
        return tr

    return mkrow
