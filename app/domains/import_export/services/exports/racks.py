import io

from ._helpers import (
    COL_ALT,
    DEVICE_COLOR,
    _rack_devices,
    _make_xlsx_helpers,
    _make_ods_mkrow,
    _to_csv,
)
from openpyxl import Workbook
from openpyxl.styles import Border, Side


def _rack_xlsx(data: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    write_header, style_row, autowidth = _make_xlsx_helpers(brd)

    ws = wb.create_sheet("Rack-Inventar")
    write_header(ws, ["Rack", "Standort", "U-von", "U-bis", "Hostname", "Typ", "IP",
                       "Phase", "Anschlussleistung (W)", "Hersteller", "Modell", "Bemerkung"])
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            u_bis = (d.get("u_pos") or 0) + (d.get("u_h") or 1) - 1
            ws.append([r["name"], r["standort"], d.get("u_pos", "–"), u_bis,
                       d["hostname"], d["typ"], d.get("ip", "–"), d.get("phase", "–"),
                       d.get("watt") or "–", d.get("hersteller", "–"),
                       d.get("modell", "–"), d.get("bemerkung", "")])
            style_row(ws, ws.max_row, fill_hex=DEVICE_COLOR.get(d.get("typ", "sonstige"), COL_ALT))
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _rack_ods(data: dict) -> bytes:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table

    doc = OpenDocumentSpreadsheet()
    mkrow = _make_ods_mkrow()
    t = Table(name="Rack-Inventar")
    t.addElement(mkrow(["Rack", "Standort", "U-von", "U-bis", "Hostname", "Typ", "IP",
                         "Phase", "Anschlussleistung (W)", "Hersteller", "Modell"]))
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            t.addElement(mkrow([
                r["name"], r["standort"], d.get("u_pos", "–"),
                (d.get("u_pos") or 0) + (d.get("u_h") or 1) - 1,
                d["hostname"], d["typ"], d.get("ip", "–"), d.get("phase", "–"),
                d.get("watt", "–"), d.get("hersteller", "–"), d.get("modell", "–"),
            ]))
    doc.spreadsheet.addElement(t)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _rack_csv(data: dict) -> bytes:
    rows = []
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            rows.append([r["name"], r["standort"], d.get("u_pos", ""),
                         (d.get("u_pos") or 0) + (d.get("u_h") or 1) - 1,
                         d["hostname"], d["typ"], d.get("ip", ""), d.get("phase", ""),
                         d.get("watt", ""), d.get("hersteller", ""), d.get("modell", ""),
                         d.get("bemerkung", "")])
    return _to_csv(["rack", "standort", "u_von", "u_bis", "hostname", "typ", "ip", "phase",
                    "anschlussleistung_w", "hersteller", "modell", "bemerkung"], rows)
