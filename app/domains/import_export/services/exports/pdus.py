import io

from ._helpers import (
    COL_ALT,
    COL_WARN,
    COL_WHITE,
    PHASE_COL,
    _dev_map,
    _rack_map,
    _pdu_outlets,
    _make_xlsx_helpers,
    _make_ods_mkrow,
    _to_csv,
)
from openpyxl import Workbook
from openpyxl.styles import Border, Side

_HEADER = [
    "PDU",
    "Rack",
    "Steckdose",
    "Phase",
    "Steckdosentyp",
    "Max W",
    "Angeschlossenes Gerät",
    "Gerät-Rack",
    "Port",
    "Schaltbar",
    "Status",
]
_HEADER_CSV = [
    "pdu",
    "rack",
    "steckdose",
    "phase",
    "steckdosentyp",
    "max_watt",
    "geraet",
    "geraet_rack",
    "port",
    "schaltbar",
    "status",
]


def _pdu_xlsx(data: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    write_header, style_row, autowidth = _make_xlsx_helpers(brd)
    dm = _dev_map(data)
    rm = _rack_map(data)

    ws = wb.create_sheet("PDU-Belegung")
    write_header(ws, _HEADER)
    for pdu in [d for d in data["devices"] if d.get("typ") == "pdu"]:
        rack_name = rm.get(pdu.get("rack_id"), {}).get("name", "–")
        outlets = _pdu_outlets(data, pdu["id"])
        if not outlets:
            ws.append(
                [
                    pdu["hostname"],
                    rack_name,
                    "–",
                    "–",
                    "–",
                    "–",
                    "–",
                    "–",
                    "–",
                    "–",
                    "kein Outlet",
                ]
            )
            style_row(ws, ws.max_row, fill_hex=COL_ALT)
            continue
        for o in outlets:
            dev = dm.get(o.get("connected_device_id"))
            dev_rack = rm.get(dev.get("rack_id") if dev else None, {}).get("name", "–")
            ws.append(
                [
                    pdu["hostname"],
                    rack_name,
                    o.get("outlet_name", "–"),
                    o.get("phase", "–"),
                    o.get("steckdosentyp", "–"),
                    o.get("max_watt", "–"),
                    dev["hostname"] if dev else "frei",
                    dev_rack,
                    o.get("connected_port", "–"),
                    "✓" if o.get("schaltbar") else "–",
                    "belegt" if dev else "frei",
                ]
            )
            fill = COL_WARN if not dev else PHASE_COL.get(o.get("phase", ""), COL_WHITE)
            style_row(ws, ws.max_row, fill_hex=fill)
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _pdu_ods(data: dict) -> bytes:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table

    doc = OpenDocumentSpreadsheet()
    mkrow = _make_ods_mkrow()
    dm = _dev_map(data)
    rm = _rack_map(data)

    t = Table(name="PDU-Belegung")
    t.addElement(mkrow(_HEADER))
    for pdu in [d for d in data["devices"] if d.get("typ") == "pdu"]:
        rn = rm.get(pdu.get("rack_id"), {}).get("name", "–")
        for o in _pdu_outlets(data, pdu["id"]):
            dev = dm.get(o.get("connected_device_id"))
            t.addElement(
                mkrow(
                    [
                        pdu["hostname"],
                        rn,
                        o.get("outlet_name", "–"),
                        o.get("phase", "–"),
                        o.get("steckdosentyp", "–"),
                        o.get("max_watt", "–"),
                        dev["hostname"] if dev else "frei",
                        rm.get(dev.get("rack_id") if dev else None, {}).get(
                            "name", "–"
                        ),
                        o.get("connected_port", "–"),
                        "JA" if o.get("schaltbar") else "NEIN",
                        "belegt" if dev else "frei",
                    ]
                )
            )
    doc.spreadsheet.addElement(t)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _pdu_csv(data: dict) -> bytes:
    dm = _dev_map(data)
    rm = _rack_map(data)
    rows = []
    for pdu in [d for d in data["devices"] if d.get("typ") == "pdu"]:
        rn = rm.get(pdu.get("rack_id"), {}).get("name", "")
        for o in _pdu_outlets(data, pdu["id"]):
            dev = dm.get(o.get("connected_device_id"))
            rows.append(
                [
                    pdu["hostname"],
                    rn,
                    o.get("outlet_name", ""),
                    o.get("phase", ""),
                    o.get("steckdosentyp", ""),
                    o.get("max_watt", ""),
                    dev["hostname"] if dev else "frei",
                    rm.get(dev.get("rack_id") if dev else None, {}).get("name", ""),
                    o.get("connected_port", ""),
                    "JA" if o.get("schaltbar") else "NEIN",
                    "belegt" if dev else "frei",
                ]
            )
    return _to_csv(_HEADER_CSV, rows)
