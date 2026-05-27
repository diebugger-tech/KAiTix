import io

from ._helpers import (
    COL_XRACK,
    COL_WHITE,
    _dev_map,
    _rack_map,
    _rack_devices,
    _ifaces_for_device,
    _cable_for_iface,
    _make_xlsx_helpers,
    _make_ods_mkrow,
    _to_csv,
)
from openpyxl import Workbook
from openpyxl.styles import Border, Side

_HEADER = [
    "Rack",
    "Gerät",
    "Port",
    "Typ",
    "MAC",
    "Verbunden mit",
    "Ziel-Port",
    "Kabel-Nr",
    "Kabel-Typ",
    "m",
    "Farbe",
    "★",
]
_HEADER_CSV = [
    "rack",
    "geraet",
    "port",
    "typ",
    "mac",
    "ziel_geraet",
    "ziel_port",
    "kabel_nr",
    "kabel_typ",
    "laenge_m",
    "farbe",
    "rack_uebergreifend",
]


def _build_rows(data, empty_marker="–", cross_marker="★"):
    dm = _dev_map(data)
    rm = _rack_map(data)
    rows = []
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            for iface in _ifaces_for_device(data, d["id"]):
                cable = _cable_for_iface(data, iface.get("kabel_id"))
                if cable:
                    is_von = cable.get("von_dev") == d["id"]
                    dst_id = cable.get("nach_dev") if is_von else cable.get("von_dev")
                    dst_dev = dm.get(dst_id)
                    dst_port = cable.get(
                        "nach_port" if is_von else "von_port", empty_marker
                    )
                    cross = dst_dev and dst_dev.get("rack_id") != r["id"]
                    if cross_marker == "★":
                        dst_rn = rm.get(
                            dst_dev.get("rack_id") if dst_dev else None, {}
                        ).get("name", "")
                        xmark = f"★ {dst_rn}" if cross else ""
                    else:
                        xmark = cross_marker if cross else ""
                    rows.append(
                        [
                            r["name"],
                            d["hostname"],
                            iface.get("port", empty_marker),
                            iface.get("typ", empty_marker),
                            iface.get("mac", empty_marker),
                            dst_dev["hostname"] if dst_dev else empty_marker,
                            dst_port,
                            cable["nr"],
                            cable["typ"],
                            cable["laenge"],
                            cable.get("farbe", empty_marker),
                            xmark,
                        ]
                    )
                else:
                    rows.append(
                        [
                            r["name"],
                            d["hostname"],
                            iface.get("port", empty_marker),
                            iface.get("typ", empty_marker),
                            iface.get("mac", empty_marker),
                            empty_marker,
                            empty_marker,
                            empty_marker,
                            empty_marker,
                            empty_marker,
                            empty_marker,
                            "",
                        ]
                    )
    return rows


def _iface_xlsx(data: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    write_header, style_row, autowidth = _make_xlsx_helpers(brd)

    ws = wb.create_sheet("Ports & Interfaces")
    write_header(ws, _HEADER)
    for row in _build_rows(data):
        ws.append(row)
        style_row(ws, ws.max_row, fill_hex=COL_XRACK if row[-1] else COL_WHITE)
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _iface_ods(data: dict) -> bytes:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table

    doc = OpenDocumentSpreadsheet()
    mkrow = _make_ods_mkrow()
    t = Table(name="Ports & Interfaces")
    t.addElement(mkrow(_HEADER))
    for row in _build_rows(data, empty_marker="–", cross_marker="★"):
        t.addElement(mkrow(row))
    doc.spreadsheet.addElement(t)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _iface_csv(data: dict) -> bytes:
    rows = _build_rows(data, empty_marker="", cross_marker="JA")
    return _to_csv(_HEADER_CSV, rows)
