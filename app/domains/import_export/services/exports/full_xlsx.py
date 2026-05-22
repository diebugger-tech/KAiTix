import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side

from ._helpers import (
    COL_XRACK,
    COL_WHITE,
    COL_ALT,
    COL_WARN,
    DEVICE_COLOR,
    PHASE_COL,
    _dev_map,
    _rack_map,
    _phase_loads,
    _rack_devices,
    _ifaces_for_device,
    _cable_for_iface,
    _pdu_outlets,
    _make_xlsx_helpers,
)


def _xlsx(data: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    write_header, style_row, autowidth = _make_xlsx_helpers(brd)

    ts = datetime.now().strftime("%d.%m.%Y %H:%M")
    dm = _dev_map(data)
    rm = _rack_map(data)

    # Übersicht
    ws = wb.create_sheet("Übersicht")
    ws["A1"] = "KAiTix — Technische Dokumentation"
    ws["A1"].font = Font(bold=True, size=14, name="Arial")
    ws["A2"] = f"Erstellt: {ts}"
    ws["A2"].font = Font(size=9, name="Arial", color="888888")
    ws.append([])
    write_header(ws, ["Rack", "Standort", "HE", "Geräte", "L1 kW", "L2 kW", "L3 kW", "Gesamt kW"])
    for r in data["racks"]:
        ph = _phase_loads(data, r["id"])
        devs = _rack_devices(data, r["id"])
        ws.append([
            r["name"], r["standort"], r["hoehe_u"], len(devs),
            round(ph["L1"] / 1000, 2), round(ph["L2"] / 1000, 2),
            round(ph["L3"] / 1000, 2), round(sum(ph.values()) / 1000, 2),
        ])
        style_row(ws, ws.max_row)
    autowidth(ws)
    ws.freeze_panes = "A2"

    # Rack-Belegung
    ws = wb.create_sheet("Rack-Belegung")
    write_header(ws, ["Rack", "Standort", "U-von", "U-bis", "Hostname", "Typ", "IP",
                       "Phase", "Anschlussleistung (W)", "Hersteller", "Modell", "Bemerkung"])
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            u_bis = (d.get("u_pos") or 0) + (d.get("u_h") or 1) - 1
            ws.append([
                r["name"], r["standort"], d.get("u_pos", "–"), u_bis,
                d["hostname"], d["typ"], d.get("ip", "–"), d.get("phase", "–"),
                d.get("watt") or "–", d.get("hersteller", "–"), d.get("modell", "–"),
                d.get("bemerkung", ""),
            ])
            style_row(ws, ws.max_row, fill_hex=DEVICE_COLOR.get(d.get("typ", "sonstige"), COL_ALT))
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Ports & Interfaces
    ws = wb.create_sheet("Ports & Interfaces")
    write_header(ws, ["Rack", "Gerät", "Port", "Typ", "MAC", "Verbunden mit",
                       "Ziel-Port", "Kabel-Nr", "Kabel-Typ", "m", "Farbe", "★"])
    for r in data["racks"]:
        for d in _rack_devices(data, r["id"]):
            for iface in _ifaces_for_device(data, d["id"]):
                cable = _cable_for_iface(data, iface.get("kabel_id"))
                if cable:
                    is_von = cable.get("von_dev") == d["id"]
                    dst_id = cable.get("nach_dev") if is_von else cable.get("von_dev")
                    dst_dev = dm.get(dst_id)
                    dst_port = cable.get("nach_port" if is_von else "von_port", "–")
                    cross = dst_dev and dst_dev.get("rack_id") != r["id"]
                    dst_rn = rm.get(dst_dev.get("rack_id") if dst_dev else None, {}).get("name", "")
                    xmark = f"★ {dst_rn}" if cross else ""
                    row = [r["name"], d["hostname"], iface.get("port", "–"), iface.get("typ", "–"),
                           iface.get("mac", "–"), dst_dev["hostname"] if dst_dev else "–",
                           dst_port, cable["nr"], cable["typ"], cable["laenge"],
                           cable.get("farbe", "–"), xmark]
                else:
                    row = [r["name"], d["hostname"], iface.get("port", "–"), iface.get("typ", "–"),
                           iface.get("mac", "–"), "–", "–", "–", "–", "–", "–", ""]
                ws.append(row)
                style_row(ws, ws.max_row, fill_hex=COL_XRACK if row[-1] else COL_WHITE)
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Kabelliste
    ws = wb.create_sheet("Kabelliste")
    write_header(ws, ["Kabel-Nr", "Typ", "Länge m", "Farbe", "Von-Gerät", "Von-Rack",
                       "Von-Port", "Nach-Gerät", "Nach-Rack", "Nach-Port",
                       "Verlegt am", "Verlegt von", "★"])
    for c in sorted(data.get("cables", []), key=lambda x: x["nr"]):
        vd = dm.get(c.get("von_dev"))
        nd = dm.get(c.get("nach_dev"))
        vr = rm.get(vd.get("rack_id") if vd else None, {}).get("name", "–")
        nr = rm.get(nd.get("rack_id") if nd else None, {}).get("name", "–")
        cross = vd and nd and vd.get("rack_id") != nd.get("rack_id")
        ws.append([c["nr"], c["typ"], c["laenge"], c.get("farbe", "–"),
                   vd["hostname"] if vd else "–", vr, c.get("von_port", "–"),
                   nd["hostname"] if nd else "–", nr, c.get("nach_port", "–"),
                   c.get("verlegt_am", "–"), c.get("verlegt_von", "–"),
                   "★ JA" if cross else ""])
        style_row(ws, ws.max_row, fill_hex=COL_XRACK if cross else COL_WHITE)
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # PDU-Belegung
    ws = wb.create_sheet("PDU-Belegung")
    write_header(ws, ["PDU", "Rack", "Steckdose", "Phase", "Steckdosentyp", "Max W",
                       "Angeschlossenes Gerät", "Gerät-Rack", "Port", "Schaltbar", "Status"])
    for pdu in [d for d in data["devices"] if d.get("typ") == "pdu"]:
        rack_name = rm.get(pdu.get("rack_id"), {}).get("name", "–")
        outlets = _pdu_outlets(data, pdu["id"])
        if not outlets:
            ws.append([pdu["hostname"], rack_name, "–", "–", "–", "–", "–", "–", "–", "–", "kein Outlet"])
            style_row(ws, ws.max_row, fill_hex=COL_ALT)
            continue
        for o in outlets:
            dev = dm.get(o.get("connected_device_id"))
            dev_rack = rm.get(dev.get("rack_id") if dev else None, {}).get("name", "–")
            ws.append([pdu["hostname"], rack_name, o.get("outlet_name", "–"), o.get("phase", "–"),
                       o.get("steckdosentyp", "–"), o.get("max_watt", "–"),
                       dev["hostname"] if dev else "frei", dev_rack,
                       o.get("connected_port", "–"), "✓" if o.get("schaltbar") else "–",
                       "belegt" if dev else "frei"])
            fill = COL_WARN if not dev else PHASE_COL.get(o.get("phase", ""), COL_WHITE)
            style_row(ws, ws.max_row, fill_hex=fill)
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Import-Vorlage
    ws = wb.create_sheet("Import-Vorlage")
    ws["A1"] = "Hardware-Import-Vorlage"
    ws["A1"].font = Font(bold=True, size=11, name="Arial", color="AA0000")
    ws.append([])
    write_header(ws, ["hostname*", "typ*", "rack_name*", "u_position*", "u_hoehe",
                       "ip_adresse", "phase", "tdp_watt", "hersteller", "modell",
                       "seriennummer", "bemerkung"], col_hex="2E4057")
    ws.append(["srv-beispiel-01", "server", "RACK-01", "40", "2",
               "192.168.1.100", "L1", "350", "Dell", "PowerEdge R650", "SN123456", "Beispielzeile"])
    style_row(ws, ws.max_row, fill_hex="FFFDE7")
    ws["A4"] = "* Pflichtfelder  |  typ: server | switch | pdu | firewall | storage | sonstige  |  phase: L1 | L2 | L3"
    ws["A4"].font = Font(size=8, name="Arial", color="888888", italic=True)
    autowidth(ws)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
