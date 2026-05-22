import io

from ._helpers import (
    COL_XRACK,
    COL_WHITE,
    _dev_map,
    _rack_map,
    _make_xlsx_helpers,
    _make_ods_mkrow,
    _to_csv,
)
from openpyxl import Workbook
from openpyxl.styles import Border, Side

_HEADER = ["Kabel-Nr", "Typ", "Länge m", "Farbe", "Von-Gerät", "Von-Rack", "Von-Port",
           "Nach-Gerät", "Nach-Rack", "Nach-Port", "Verlegt am", "Verlegt von", "★"]
_HEADER_CSV = ["kabel_nr", "typ", "laenge_m", "farbe", "von_geraet", "von_rack", "von_port",
               "nach_geraet", "nach_rack", "nach_port", "verlegt_am", "verlegt_von", "rack_uebergreifend"]


def _cable_xlsx(data: dict) -> bytes:
    wb = Workbook()
    wb.remove(wb.active)
    thin = Side(style="thin", color="CCCCCC")
    brd = Border(left=thin, right=thin, top=thin, bottom=thin)
    write_header, style_row, autowidth = _make_xlsx_helpers(brd)
    dm = _dev_map(data)
    rm = _rack_map(data)

    ws = wb.create_sheet("Kabelliste")
    write_header(ws, _HEADER)
    for c in sorted(data.get("cables", []), key=lambda x: x["nr"]):
        vd = dm.get(c.get("von_dev"))
        nd = dm.get(c.get("nach_dev"))
        vr = rm.get(vd.get("rack_id") if vd else None, {}).get("name", "–")
        nr = rm.get(nd.get("rack_id") if nd else None, {}).get("name", "–")
        cross = vd and nd and vd.get("rack_id") != nd.get("rack_id")
        ws.append([c["nr"], c["typ"], c["laenge"], c.get("farbe", "–"),
                   vd["hostname"] if vd else "–", vr, c.get("von_port", "–"),
                   nd["hostname"] if nd else "–", nr, c.get("nach_port", "–"),
                   c.get("verlegt_am", "–"), c.get("verlegt_von", "–"),
                   "★ JA" if cross else ""])
        style_row(ws, ws.max_row, fill_hex=COL_XRACK if cross else COL_WHITE)
    autowidth(ws)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cable_ods(data: dict) -> bytes:
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table

    doc = OpenDocumentSpreadsheet()
    mkrow = _make_ods_mkrow()
    dm = _dev_map(data)
    rm = _rack_map(data)

    t = Table(name="Kabelliste")
    t.addElement(mkrow(["Kabel-Nr", "Typ", "m", "Farbe", "Von-Gerät", "Von-Rack", "Von-Port",
                         "Nach-Gerät", "Nach-Rack", "Nach-Port", "★"]))
    for c in sorted(data.get("cables", []), key=lambda x: x["nr"]):
        vd = dm.get(c.get("von_dev"))
        nd = dm.get(c.get("nach_dev"))
        vr = rm.get(vd.get("rack_id") if vd else None, {}).get("name", "–")
        nr = rm.get(nd.get("rack_id") if nd else None, {}).get("name", "–")
        cross = "★" if (vd and nd and vd.get("rack_id") != nd.get("rack_id")) else ""
        t.addElement(mkrow([c["nr"], c["typ"], c["laenge"], c.get("farbe", "–"),
                            vd["hostname"] if vd else "–", vr, c.get("von_port", "–"),
                            nd["hostname"] if nd else "–", nr, c.get("nach_port", "–"), cross]))
    doc.spreadsheet.addElement(t)
    buf = io.BytesIO()
    doc.save(buf)
    return buf.getvalue()


def _cable_csv(data: dict) -> bytes:
    dm = _dev_map(data)
    rm = _rack_map(data)
    rows = []
    for c in sorted(data.get("cables", []), key=lambda x: x["nr"]):
        vd = dm.get(c.get("von_dev"))
        nd = dm.get(c.get("nach_dev"))
        vr = rm.get(vd.get("rack_id") if vd else None, {}).get("name", "")
        nr = rm.get(nd.get("rack_id") if nd else None, {}).get("name", "")
        cross = "JA" if (vd and nd and vd.get("rack_id") != nd.get("rack_id")) else ""
        rows.append([c["nr"], c["typ"], c["laenge"], c.get("farbe", ""),
                     vd["hostname"] if vd else "", vr, c.get("von_port", ""),
                     nd["hostname"] if nd else "", nr, c.get("nach_port", ""),
                     c.get("verlegt_am", ""), c.get("verlegt_von", ""), cross])
    return _to_csv(_HEADER_CSV, rows)
